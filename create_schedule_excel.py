import datetime
import json

from openpyxl import Workbook, load_workbook

from database import execute_select_query

with open('config/db_config.json', "r") as json_file:
    config_json = json.load(json_file)
    SCHEMA = config_json['schema']

CASH_SCHEDULE = f"select DISTINCT schedule_group from {SCHEMA}.rd_account_transactions where rd_date = '(date)' and is_cash = True and schedule_number is NULL and schedule_group is NOT NULL;"
CHEQUE_SCHEDULE = f"select DISTINCT schedule_group from {SCHEMA}.rd_account_transactions where rd_date = '(date)' and is_cash = False and schedule_number is NULL and schedule_group is NOT NULL;"

CASH_SCHEDULE_DETAILS = "select " \
                        "t.account_no," \
                        "m.investor_name," \
                        "t.no_of_installments," \
                        "t.no_of_installments * m.denomination," \
                        "t.rd_date," \
                        "case " \
                        "when m.is_extended = true then m.new_card_number else card_number end as card_number," \
                        "m.total_months_paid," \
                        "m.last_deposit_date," \
                        "m.next_installment_date" \
                        " from " \
                        f"{SCHEMA}.rd_account_transactions t, {SCHEMA}.rd_master m " \
                        "where t.account_no = m.account_no and " \
                        "t.schedule_group = (scheduleGroup) and " \
                        "t.is_cash = True and t.schedule_number IS NULL " \
                        "ORDER BY t.account_no;"

CHEQUE_SCHEDULE_DETAILS = "select " \
                          "t.account_no," \
                          "m.investor_name," \
                          "t.no_of_installments," \
                          "t.no_of_installments * m.denomination," \
                          "t.rd_date," \
                          "case " \
                          "when m.is_extended = true then m.new_card_number else card_number end as card_number," \
                          "t.cheque_number," \
                          "m.bank_account_no," \
                          "m.total_months_paid," \
                          "m.last_deposit_date," \
                          "m.next_installment_date" \
                          " from " \
                          f"{SCHEMA}.rd_account_transactions t, {SCHEMA}.rd_master m " \
                          "where t.account_no = m.account_no and " \
                          "t.schedule_group = (scheduleGroup) and " \
                          "t.is_cash = False and t.schedule_number IS NULL " \
                          "ORDER BY t.cheque_number;"

filename = None


def update_file_name(file):
    global filename
    filename = file


def get_file_name():
    global filename
    return filename


def create_workbook(type, date):
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    create_header(ws, type)
    file_name_str = f'{str(date)}_{type}.xlsx'
    update_file_name(file_name_str)
    wb.save(file_name_str)


def create_header(ws, type):
    if type == 'cash':
        ws.append(("Account_No", "Name", "No Of Installment", "Amount", "RD Date", "Card Number",
                   "MonthsPaid", "Last Date", "Next Date"))
    else:
        ws.append(("Account_No", "Name", "No Of Installment",  "Amount","RD Date", "Card Number", "Cheque Number",
                   "Account Number", "MonthsPaid", "Last Date", "Next Date"))


def create_sheet(sheet_number, schedule_transactions):
    type = 'cash' if len(schedule_transactions) == 9 else 'cheque'
    wb = load_workbook(get_file_name())
    sheet_name = f"Sheet{str(sheet_number)}"
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.create_sheet(sheet_name)
        create_header(ws, type)
    if type == 'cheque':
        ws.append(
            (schedule_transactions[0], schedule_transactions[1], schedule_transactions[2],
             float(schedule_transactions[3]),
             schedule_transactions[4], schedule_transactions[5], schedule_transactions[6], schedule_transactions[7],
             schedule_transactions[8], schedule_transactions[9], schedule_transactions[10]))
    else:
        ws.append(
            (schedule_transactions[0], schedule_transactions[1], schedule_transactions[2],
             float(schedule_transactions[3]),
             schedule_transactions[4], schedule_transactions[5], schedule_transactions[6], schedule_transactions[7],
             schedule_transactions[8]))
    wb.save(get_file_name())


def create_cash_schedules(date):
    statement = CASH_SCHEDULE.replace("(date)", str(date))
    total_schedules = execute_select_query(statement)
    output = [item for t in total_schedules for item in t]
    output.sort()
    for i in output:
        statement = CASH_SCHEDULE_DETAILS.replace("(scheduleGroup)", f"{i}")
        accounts_in_schedule = execute_select_query(statement)
        for schedule in accounts_in_schedule:
            create_sheet(i, schedule)


def create_cheque_schedules(date):
    statement = CHEQUE_SCHEDULE.replace("(date)", str(date))
    total_schedules = execute_select_query(statement)
    output = [item for t in total_schedules for item in t]
    output.sort()
    for i in output:
        statement = CHEQUE_SCHEDULE_DETAILS.replace("(scheduleGroup)", f"{i}")
        accounts_in_schedule = execute_select_query(statement)
        for schedule in accounts_in_schedule:
            create_sheet(i, schedule)


def finish_workbook():
    wb = load_workbook(get_file_name())
    for sheet in wb.worksheets:
        max_row = sheet.max_row
        sheet.cell(row=max_row + 2, column=4).value = f'=SUM(D{2}:D{max_row})'
    wb.save(get_file_name())


if __name__ == '__main__':
    today_date = datetime.date.today()
    if today_date.day <= 15:
        date = datetime.date(today_date.year, today_date.month, 1)
        create_workbook("cash", date)
        create_cash_schedules(date)
        finish_workbook()
        create_workbook("cheque", date)
        create_cheque_schedules(date)
        finish_workbook()
    else:
        date = datetime.date(today_date.year, today_date.month, 16)
        create_workbook("cash", date)
        create_cash_schedules(date)
        finish_workbook()
        create_workbook("cheque", date)
        create_cheque_schedules(date)
        finish_workbook()
