import json
import datetime

from openpyxl import load_workbook

from database import execute_insert_query

with open('config/db_config.json', "r") as json_file:
    config_json = json.load(json_file)
    file_name = config_json['file_name']


def enter_into_transaction_table(insert_dict):
    execute_insert_query("transaction", insert_dict)


if __name__ == '__main__':
    wb = load_workbook(file_name)
    ws = wb['Sheet1']
    today_date = datetime.date.today()
    if today_date.day <= 15:
        date = datetime.date(today_date.year, today_date.month, 1)
    else:
        date = datetime.date(today_date.year, today_date.month, 16)

    for row in ws.iter_rows(min_row=2):
        insert_dict = {"account_no": row[0].value, "rd_date": str(date), "schedule_group": (int(row[4].value)),
                       "is_cash": bool(row[5].value), "no_of_installments": int(row[2].value)}
        if row[6].value:
            insert_dict['cheque_number'] = row[6].value
            insert_dict['is_cash'] = False
        print(insert_dict)
        enter_into_transaction_table(insert_dict)
