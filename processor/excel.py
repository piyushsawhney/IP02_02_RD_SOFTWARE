import datetime

from openpyxl import Workbook, load_workbook

from processor.schedule import ScheduleProcessor

filename = None


def update_file_name(file):
    global filename
    filename = file


def get_file_name():
    global filename
    return filename


class ExcelProcessor:
    @staticmethod
    def adjust_column_size(ws):
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column].width = adjusted_width

    @staticmethod
    def create_file(is_cash, rd_date):
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ExcelProcessor.create_header(ws, is_cash)
        file_name_str = f"files\\{str(rd_date)}_{'cash' if is_cash else 'cheque'}.xlsx"
        update_file_name(file_name_str)
        ExcelProcessor.adjust_column_size(ws)
        wb.save(file_name_str)

    @staticmethod
    def create_header(ws, is_cash):
        if is_cash:
            ws.append(("Account Number", "Client Name", "No. Of Installment", "Amount", "RD Date",
                       "Months Paid", "Last Paid Date", "Next Due Date"))
        else:
            ws.append(("Account Number", "Client Name", "No. Of Installment", "Amount", "RD Date", "Cheque Number",
                       "Savings Account Number", "Months Paid", "Last Paid Date", "Next Due Date"))

    @staticmethod
    def update_file(is_cash, schedule_group, transaction_tupple):
        wb = load_workbook(get_file_name())
        sheet_name = f"Sheet{str(schedule_group[0])}"
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            ws = wb.create_sheet(sheet_name)
            ExcelProcessor.create_header(ws, is_cash)
        ws.append(transaction_tupple)
        ExcelProcessor.adjust_column_size(ws)
        wb.save(get_file_name())

    @staticmethod
    def update_total():
        wb = load_workbook(get_file_name())
        for sheet in wb.worksheets:
            max_row = sheet.max_row
            sheet.cell(row=max_row + 2, column=4).value = f'=SUM(D{2}:D{max_row})'
        wb.save(get_file_name())


def process_schedules(is_cash, rd_date):
    ExcelProcessor.create_file(is_cash, rd_date)
    for schedule_group in ScheduleProcessor.get_schedule_list(is_cash, rd_date):
        transaction_list = ScheduleProcessor.get_schedule_details(is_cash, rd_date, schedule_group)
        for transaction in transaction_list:
            ExcelProcessor.update_file(is_cash, schedule_group, transaction)
    ExcelProcessor.update_total()


if __name__ == '__main__':
    today_date = datetime.date.today()
    if today_date.day <= 15:
        rd_date = datetime.date(today_date.year, today_date.month, 1)
    else:
        rd_date = datetime.date(today_date.year, today_date.month, 16)
    for is_cash in [True, False]:
        process_schedules(is_cash, rd_date)
