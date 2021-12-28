import datetime
import json

from openpyxl import load_workbook

from processor.app_config.db_sql import ADD_ACCOUNT_TRANSACTIONS, UPDATE_RD_CLIENTS
from processor.db.database import execute_statement


class TransactionEntry:
    @staticmethod
    def read_from_file(date):
        with open('configuration/configuration.json', "r") as json_file:
            config_json = json.load(json_file)
            file_name = config_json['file_name']
        wb = load_workbook(file_name)
        ws = wb['Sheet1']
        for row in ws.iter_rows(min_row=2):
            # SQL: rd_date, account_no, is_cash, no_of_installments, schedule_group, cheque_number
            # EXCEL: AccountNumber, Name, NumberOfInstallement, Amount, ScheduleNumber, Cash, ChequeNumber
            values = (str(date), row[0].value, bool(row[5].value), int(row[2].value), int(row[4].value),
                      row[6].value if row[6].value else None,)
            TransactionEntry.update_to_db(values)

    @staticmethod
    def read_from_file_card(date):
        with open('configuration/configuration.json', "r") as json_file:
            config_json = json.load(json_file)
            file_name = config_json['file_name']
        wb = load_workbook(file_name)
        ws = wb['Sheet1']
        for row in ws.iter_rows(min_row=2):
            # SQL: rd_date, account_no, is_cash, no_of_installments, schedule_group, cheque_number
            # EXCEL: AccountNumber, Name, NumberOfInstallement, Amount, ScheduleNumber, Cash, ChequeNumber
            # values = (str(date), row[0].value, bool(row[5].value), int(row[2].value), int(row[4].value),
            #           row[6].value if row[6].value else None,)
            values = (row[8].value if row[8].value else None, row[8].value if row[8].value else None,
                      row[7].value if row[7].value else None, row[0].value)
            execute_statement(UPDATE_RD_CLIENTS,values)

    @staticmethod
    def update_to_db(values):
        execute_statement(ADD_ACCOUNT_TRANSACTIONS, values)


if __name__ == '__main__':
    today_date = datetime.date.today()
    if today_date.day <= 15:
        date = datetime.date(today_date.year, today_date.month, 1)
    else:
        date = datetime.date(today_date.year, today_date.month, 16)
    TransactionEntry.read_from_file(date)
